# -*- coding: utf-8 -*-
"""解析 AI 竞赛·证书综合映射 Excel"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _split_list(text: str) -> list[str]:
    if not text:
        return []
    return [p.strip() for p in re.split(r"[,，、/;；\n]+", text) if p.strip()]


def _sheet_rows(wb: openpyxl.Workbook, name: str) -> list[list[Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    return [list(row) if row else [] for row in ws.iter_rows(values_only=True)]


def _find_sheet(wb: openpyxl.Workbook, test: str) -> str | None:
    for n in wb.sheetnames:
        if test in n:
            return n
    return None


def _find_header_row(rows: list[list[Any]], markers: list[str]) -> int:
    for i, row in enumerate(rows[:15]):
        text = " ".join(_cell_str(c) for c in row)
        if all(m in text for m in markers):
            return i
    return 0


def _parse_competitions(wb: openpyxl.Workbook) -> list[dict]:
    sn = _find_sheet(wb, "竞赛总览")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    hi = _find_header_row(rows, ["序号", "竞赛名称"])
    headers = [_cell_str(c) for c in rows[hi]]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: list[dict] = []
    for row in rows[hi + 1 :]:
        name = _cell_str(row[idx.get("竞赛名称", 1)] if len(row) > 1 else "")
        if not name:
            continue
        out.append({
            "id": len(out) + 1,
            "name": name,
            "category": _cell_str(row[idx.get("竞赛类别", 2)] if len(row) > 2 else ""),
            "organizer": _cell_str(row[idx.get("主办/承办单位", 3)] if len(row) > 3 else ""),
            "level": _cell_str(row[idx.get("竞赛级别", 4)] if len(row) > 4 else ""),
            "official": _cell_str(row[idx.get("官方/非官方", 5)] if len(row) > 5 else ""),
            "rating": _cell_str(row[idx.get("价值评级", 6)] if len(row) > 6 else ""),
            "labs": _split_list(_cell_str(row[idx.get("对应实训室", 7)] if len(row) > 7 else "")),
            "jobs": _split_list(_cell_str(row[idx.get("对应岗位", 8)] if len(row) > 8 else "")),
            "summary": _cell_str(row[idx.get("竞赛简介", 9)] if len(row) > 9 else ""),
            "audience": _cell_str(row[idx.get("适用层次", 10)] if len(row) > 10 else ""),
        })
    return out


def _parse_lab_competitions(wb: openpyxl.Workbook) -> list[dict]:
    sn = _find_sheet(wb, "实训室-竞赛")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    hi = _find_header_row(rows, ["实训室名称"])
    out: list[dict] = []
    for row in rows[hi + 1 :]:
        lab = _cell_str(row[1] if len(row) > 1 else "")
        if not lab or lab.startswith("【"):
            continue
        count_val = row[2] if len(row) > 2 else 0
        try:
            count = int(count_val) if count_val is not None else 0
        except (TypeError, ValueError):
            count = 0
        out.append({
            "lab": lab,
            "count": count,
            "competitionText": _cell_str(row[3] if len(row) > 3 else ""),
            "audience": _cell_str(row[4] if len(row) > 4 else ""),
        })
    return out


def _parse_job_competitions(wb: openpyxl.Workbook) -> list[dict]:
    sn = _find_sheet(wb, "岗位-竞赛")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    hi = _find_header_row(rows, ["岗位"])
    out: list[dict] = []
    for row in rows[hi + 1 :]:
        label = _cell_str(row[1] if len(row) > 1 else "")
        if not label:
            continue
        is_category = label.startswith("【") and label.endswith("】")
        if is_category:
            out.append({"label": label, "isCategory": True, "count": 0, "competitionText": "", "audience": ""})
            continue
        count_val = row[2] if len(row) > 2 else 0
        try:
            count = int(count_val) if count_val is not None else 0
        except (TypeError, ValueError):
            count = 0
        out.append({
            "label": label,
            "isCategory": False,
            "count": count,
            "competitionText": _cell_str(row[3] if len(row) > 3 else ""),
            "audience": _cell_str(row[4] if len(row) > 4 else ""),
        })
    return out


def _parse_category_stats(wb: openpyxl.Workbook) -> list[dict]:
    sn = _find_sheet(wb, "竞赛分类统计")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    hi = _find_header_row(rows, ["竞赛类别"])
    out: list[dict] = []
    for row in rows[hi + 1 :]:
        cat = _cell_str(row[0] if row else "")
        if not cat:
            continue
        def _num(i: int) -> int:
            try:
                v = row[i] if len(row) > i else 0
                return int(v) if v is not None else 0
            except (TypeError, ValueError):
                return 0
        out.append({
            "category": cat,
            "count": _num(1),
            "star5": _num(2),
            "star4": _num(3),
            "star3": _num(4),
            "examples": _cell_str(row[5] if len(row) > 5 else ""),
            "audience": _cell_str(row[6] if len(row) > 6 else ""),
        })
    return out


def _parse_certificates(wb: openpyxl.Workbook) -> list[dict]:
    sn = _find_sheet(wb, "证书总表")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    hi = _find_header_row(rows, ["序号", "证书"])
    headers = [_cell_str(c) for c in rows[hi]]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: list[dict] = []
    for row in rows[hi + 1 :]:
        full = _cell_str(row[idx.get("证书全称", 3)] if len(row) > 3 else "")
        short = _cell_str(row[idx.get("证书简称", 4)] if len(row) > 4 else "")
        if not full and not short:
            continue
        out.append({
            "id": len(out) + 1,
            "category": _cell_str(row[idx.get("认证类别", 1)] if len(row) > 1 else ""),
            "org": _cell_str(row[idx.get("认证机构", 2)] if len(row) > 2 else ""),
            "fullName": full,
            "shortName": short,
            "level": _cell_str(row[idx.get("认证等级/级别", 5)] if len(row) > 5 else ""),
            "field": _cell_str(row[idx.get("认证方向/领域", 6)] if len(row) > 6 else ""),
            "labs": _split_list(_cell_str(row[idx.get("适用实训室", 7)] if len(row) > 7 else "")),
            "examForm": _cell_str(row[idx.get("考试形式", 8)] if len(row) > 8 else ""),
            "recognition": _cell_str(row[idx.get("含金量/行业认可", 9)] if len(row) > 9 else ""),
            "note": _cell_str(row[idx.get("备注", 10)] if len(row) > 10 else ""),
        })
    return out


def _parse_recommendations(wb: openpyxl.Workbook) -> list[str]:
    sn = _find_sheet(wb, "参赛建议")
    if not sn:
        return []
    rows = _sheet_rows(wb, sn)
    lines: list[str] = []
    for row in rows:
        text = _cell_str(row[0] if row else "")
        if text and len(text) > 4:
            lines.append(text)
    return lines[:40]


def parse_comp_cert_workbook(filepath: str | Path) -> dict:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        competitions = _parse_competitions(wb)
        certificates = _parse_certificates(wb)
        lab_competitions = _parse_lab_competitions(wb)
        job_competitions = _parse_job_competitions(wb)
        category_stats = _parse_category_stats(wb)
        recommendations = _parse_recommendations(wb)
        return {
            "meta": {
                "total_competitions": len(competitions),
                "total_certificates": len(certificates),
                "total_labs_mapped": len(lab_competitions),
                "source_file": Path(filepath).name,
            },
            "competitions": competitions,
            "certificates": certificates,
            "labCompetitions": lab_competitions,
            "jobCompetitions": job_competitions,
            "categoryStats": category_stats,
            "recommendations": recommendations,
        }
    finally:
        wb.close()
