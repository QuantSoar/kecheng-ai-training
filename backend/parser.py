# -*- coding: utf-8 -*-
"""解析 24个实训室课程体系建设调研表（整理版）.xlsx"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from typing import Any

import openpyxl


TYPE_MAP = {
    "通识课程": "general",
    "专业基础课": "foundation",
    "专业核心课": "core",
    "专业实践课": "practice",
}

TYPE_LABELS = {v: k for k, v in TYPE_MAP.items()}

VENDOR_ALIASES = {
    "华清": "华清智言",
}


def _normalize_vendor_name(name: str) -> str:
    name = name.strip()
    return VENDOR_ALIASES.get(name, name)


def _normalize_source(source: str) -> str:
    if not source:
        return source
    return re.sub(r"厂商课程/华清(?!智言)", "厂商课程/华清智言", source)


def _merge_vendor_distribution(vendors: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for v in vendors:
        name = _normalize_vendor_name(v["vendor"])
        if name not in merged:
            merged[name] = {"vendor": name, "count": 0, "ratio": 0.0, "lab_count": 0}
        merged[name]["count"] += int(v["count"])
        merged[name]["lab_count"] = max(merged[name]["lab_count"], int(v.get("lab_count") or 0))
    total = sum(x["count"] for x in merged.values())
    result = []
    for x in merged.values():
        x["ratio"] = x["count"] / total if total else 0
        result.append(x)
    result.sort(key=lambda x: -x["count"])
    return result


def _extract_vendor_from_source(source: str) -> str:
    m = re.match(r"厂商课程/(.+)", source)
    if m:
        return _normalize_vendor_name(m.group(1).strip())
    return ""


def _apply_vendor_normalization(
    courses: list[dict], vendor_courses: list[dict], stats: dict
) -> None:
    for c in courses:
        c["source"] = _normalize_source(c.get("source", ""))
        if c.get("vendor"):
            c["vendor"] = _normalize_vendor_name(c["vendor"])
        elif c["source"]:
            extracted = _extract_vendor_from_source(c["source"])
            if extracted:
                c["vendor"] = extracted
        vd = c.get("vendor_detail")
        if vd and vd.get("vendor"):
            vd["vendor"] = _normalize_vendor_name(vd["vendor"])
    for v in vendor_courses:
        v["vendor"] = _normalize_vendor_name(v.get("vendor", ""))


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_vendor_source(source: str) -> bool:
    if not source:
        return False
    keywords = ("厂商", "华为", "软通", "讯飞", "科大讯飞", "优必选", "乐聚", "智元", "星海图", "华清智言", "华清", "AIGC厂商", "国投智能")
    return any(k in source for k in keywords)


def _parse_hours(val: Any) -> int | str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return int(val) if val == int(val) else int(val)
    s = str(val).strip()
    if re.match(r"^\d+$", s):
        return int(s)
    return s


def _find_header_row(rows: list[tuple], keywords: tuple[str, ...]) -> int:
    for i, row in enumerate(rows):
        texts = [_cell_str(c) for c in row]
        if sum(1 for t in texts if t) >= 3 and any(k in t for t in texts for k in keywords):
            return i
    return 0


def _build_vendor_distribution(vendor_courses: list[dict]) -> list[dict]:
    counts: Counter[str] = Counter()
    labs_by_vendor: dict[str, set[str]] = defaultdict(set)
    for v in vendor_courses:
        name = (v.get("vendor") or "").strip()
        if not name:
            continue
        counts[name] += 1
        if v.get("lab_name"):
            labs_by_vendor[name].add(v["lab_name"])
    total = sum(counts.values()) or 1
    result = [
        {
            "vendor": name,
            "count": cnt,
            "ratio": cnt / total,
            "lab_count": len(labs_by_vendor.get(name, set())),
        }
        for name, cnt in counts.items()
    ]
    result.sort(key=lambda x: -x["count"])
    return result


def _build_type_distribution(courses: list[dict]) -> list[dict]:
    labels = {v: k for k, v in TYPE_MAP.items()}
    counts: Counter[str] = Counter(c["type"] for c in courses)
    total = sum(counts.values()) or 1
    order = ["general", "foundation", "core", "practice", "other"]
    result = []
    for t in order:
        if counts[t]:
            result.append(
                {
                    "label": labels.get(t, t),
                    "type": t,
                    "count": counts[t],
                    "ratio": counts[t] / total,
                }
            )
    return result


def parse_workbook(filepath: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    lab_mapping = _parse_lab_names(wb)
    courses = _parse_all_courses(wb)
    vendor_courses = _parse_vendor_courses(wb)
    overview = _parse_overview(wb)
    stats = _parse_stats(wb)

    _merge_vendor_into_courses(courses, vendor_courses)
    _apply_vendor_normalization(courses, vendor_courses, stats)
    shared = _build_shared_courses(courses)
    labs = _build_labs(courses, overview, lab_mapping)

    # 以实际解析数据为准重新计算统计，避免 Excel 统计表与明细不一致
    type_stats = _build_type_distribution(courses)
    vendor_stats = _build_vendor_distribution(vendor_courses)

    return {
        "meta": {
            "total_courses": len(courses),
            "total_labs": len(labs),
            "total_vendor_courses": len(vendor_courses),
            "vendor_count": len({v["vendor"] for v in vendor_courses if v.get("vendor")}),
            "shared_course_count": len(shared),
        },
        "labs": labs,
        "courses": courses,
        "vendor_courses": vendor_courses,
        "overview": overview,
        "stats": {
            "type_distribution": type_stats,
            "vendor_distribution": vendor_stats,
        },
        "shared_courses": shared,
        "lab_mapping": lab_mapping,
    }


def _parse_lab_names(wb) -> list[dict]:
    if "实训室名称" not in wb.sheetnames:
        return []
    ws = wb["实训室名称"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows, ("实训室名称", "序号"))
    headers = [_cell_str(c) for c in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1 :]:
        if not any(row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        name = _cell_str(item.get("实训室名称"))
        if not name:
            continue
        result.append(
            {
                "id": item.get("序号"),
                "name": name,
                "original_name": _cell_str(item.get("原Excel中名称")),
            }
        )
    return result


def _parse_all_courses(wb) -> list[dict]:
    ws = wb["全部课程明细"]
    rows = list(ws.iter_rows(values_only=True))
    courses: list[dict] = []
    seq = 0

    for row in rows[4:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue

        first = row[0]
        # 分组标题行: "1  人工智能应用开发实训室"
        if isinstance(first, str) and "实训室" in first and row[1] is None:
            continue

        lab = _cell_str(row[1]) if len(row) > 1 else ""
        name = _cell_str(row[3]) if len(row) > 3 else ""
        if not lab or not name:
            continue

        seq += 1
        course_type_cn = _cell_str(row[2]) if len(row) > 2 else ""
        source = _normalize_source(_cell_str(row[7]) if len(row) > 7 else "")

        courses.append(
            {
                "id": seq,
                "lab_name": lab,
                "type": TYPE_MAP.get(course_type_cn, "other"),
                "type_label": course_type_cn or "其他",
                "name": name,
                "hours": _parse_hours(row[4]) if len(row) > 4 else None,
                "textbook": _cell_str(row[5]) if len(row) > 5 else "",
                "platform": _cell_str(row[6]) if len(row) > 6 else "",
                "source": source,
                "note": _cell_str(row[8]) if len(row) > 8 else "",
                "is_vendor": _is_vendor_source(source),
            }
        )

    return courses


def _parse_vendor_courses(wb) -> list[dict]:
    if "各个厂商课程体系" not in wb.sheetnames:
        return []
    ws = wb["各个厂商课程体系"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows, ("课程名称", "所属厂商"))
    headers = [_cell_str(c) for c in rows[header_idx]]
    result = []
    vid = 0

    for row in rows[header_idx + 1 :]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        name = _cell_str(item.get("课程名称"))
        if not name:
            continue
        vid += 1
        result.append(
            {
                "id": vid,
                "lab_name": _cell_str(item.get("归属实训室（标准）")),
                "original_lab_name": _cell_str(item.get("原始实训室名称")),
                "name": name,
                "description": _cell_str(item.get("课程描述")),
                "type": TYPE_MAP.get(_cell_str(item.get("课程类型（标准）")), "other"),
                "type_label": _cell_str(item.get("课程类型（标准）")),
                "original_type": _cell_str(item.get("原始课程类型")),
                "hardware_bound": _cell_str(item.get("是否绑定硬件")) == "是",
                "hardware": _cell_str(item.get("配套硬件")),
                "hours_total": _parse_hours(item.get("参考课时（总）")),
                "hours_theory": _parse_hours(item.get("参考课时（理论）")),
                "hours_practice": _parse_hours(item.get("参考课时（实训）")),
                "vendor": _normalize_vendor_name(_cell_str(item.get("所属厂商"))),
                "capability": _cell_str(item.get("培养能力")),
                "jobs": _cell_str(item.get("岗位辐射")),
            }
        )
    return result


def _parse_overview(wb) -> list[dict]:
    if "课程体系总览" not in wb.sheetnames:
        return []
    ws = wb["课程体系总览"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows, ("实训室名称", "课程总数"))
    headers = [_cell_str(c) for c in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1 :]:
        if not any(row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        lab = _cell_str(item.get("实训室名称"))
        if not lab or lab in ("合计", "24个实训室"):
            continue
        result.append(
            {
                "id": item.get("序号"),
                "lab_name": lab,
                "general": int(item.get("通识课程") or 0),
                "foundation": int(item.get("专业基础课") or 0),
                "core": int(item.get("专业核心课") or 0),
                "practice": int(item.get("专业实践课") or 0),
                "total": int(item.get("课程总数") or 0),
                "note": _cell_str(item.get("备注")),
            }
        )
    return result


def _parse_stats(wb) -> dict:
    stats: dict[str, Any] = {"type_distribution": [], "vendor_distribution": []}
    if "统计汇总" not in wb.sheetnames:
        return stats
    ws = wb["统计汇总"]
    rows = list(ws.iter_rows(values_only=True))

    section = None
    for row in rows:
        texts = [_cell_str(c) for c in row]
        joined = " ".join(texts)
        if "课程类型分布" in joined:
            section = "type"
            continue
        if "各实训室课程数量" in joined:
            section = "labs"
            continue
        if "厂商课程分布" in joined:
            section = "vendor"
            continue

        if section == "type":
            label = texts[0]
            if label in TYPE_MAP or label == "合计":
                count = row[1]
                ratio = row[2]
                if label != "合计" and count:
                    stats["type_distribution"].append(
                        {
                            "label": label,
                            "type": TYPE_MAP.get(label, "other"),
                            "count": int(count),
                            "ratio": float(ratio) if ratio else 0,
                        }
                    )

        if section == "vendor":
            vendor = texts[0]
            if vendor and vendor not in ("所属厂商", "合计") and row[1]:
                stats["vendor_distribution"].append(
                    {
                        "vendor": _normalize_vendor_name(vendor),
                        "count": int(row[1]),
                        "ratio": float(row[2]) if row[2] else 0,
                        "lab_count": int(row[3]) if len(row) > 3 and row[3] else 0,
                    }
                )

    stats["vendor_distribution"] = _merge_vendor_distribution(
        stats.get("vendor_distribution", [])
    )
    return stats


def _merge_vendor_into_courses(courses: list[dict], vendor_courses: list[dict]) -> None:
    index: dict[tuple[str, str], dict] = {}
    for v in vendor_courses:
        key = (v["lab_name"], v["name"])
        index[key] = v

    for c in courses:
        v = index.get((c["lab_name"], c["name"]))
        if v:
            c["vendor_detail"] = v
            c["is_vendor"] = True
            if v.get("description"):
                c["description"] = v["description"]
            if v.get("hardware"):
                c["hardware"] = v["hardware"]
            if v.get("vendor"):
                c["vendor"] = v["vendor"]


def _build_shared_courses(courses: list[dict]) -> list[dict]:
    by_name: dict[str, list[str]] = {}
    for c in courses:
        by_name.setdefault(c["name"], []).append(c["lab_name"])

    shared = []
    for name, labs in by_name.items():
        unique_labs = sorted(set(labs))
        if len(unique_labs) >= 2:
            shared.append(
                {
                    "name": name,
                    "lab_count": len(unique_labs),
                    "labs": unique_labs,
                    "course_ids": [c["id"] for c in courses if c["name"] == name],
                }
            )
    shared.sort(key=lambda x: -x["lab_count"])
    return shared


def _build_labs(courses: list[dict], overview: list[dict], lab_mapping: list[dict]) -> list[dict]:
    overview_by_name = {o["lab_name"]: o for o in overview}
    mapping_by_name = {m["name"]: m for m in lab_mapping}

    lab_names = sorted({c["lab_name"] for c in courses})
    labs = []
    for i, name in enumerate(lab_names, start=1):
        lab_courses = [c for c in courses if c["lab_name"] == name]
        ov = overview_by_name.get(name, {})
        mp = mapping_by_name.get(name, {})

        counts = {
            "general": sum(1 for c in lab_courses if c["type"] == "general"),
            "foundation": sum(1 for c in lab_courses if c["type"] == "foundation"),
            "core": sum(1 for c in lab_courses if c["type"] == "core"),
            "practice": sum(1 for c in lab_courses if c["type"] == "practice"),
            "total": len(lab_courses),
        }
        if ov:
            counts = {
                "general": ov.get("general", counts["general"]),
                "foundation": ov.get("foundation", counts["foundation"]),
                "core": ov.get("core", counts["core"]),
                "practice": ov.get("practice", counts["practice"]),
                "total": ov.get("total", counts["total"]),
            }

        labs.append(
            {
                "id": ov.get("id") or mp.get("id") or i,
                "name": name,
                "original_name": mp.get("original_name", ""),
                "course_counts": counts,
                "vendor_course_count": sum(1 for c in lab_courses if c["is_vendor"]),
                "course_ids": [c["id"] for c in lab_courses],
            }
        )

    labs.sort(key=lambda x: x["course_counts"]["total"], reverse=True)
    return labs
