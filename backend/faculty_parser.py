# -*- coding: utf-8 -*-
"""解析 师资多维分析表.xlsx"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

HEADER_ROW = 4  # 1-based: 第4行是列名
DATA_START = 5

MAIN_COLUMNS = [
    "seq", "name", "gender", "education", "university", "major", "title_level",
    "source", "title_dim", "level", "experience", "cert_category", "availability",
    "ability", "field", "keywords", "matched_labs", "matched_lab_count",
    "teaching_mgmt", "cooperation_mode", "fee_level", "priority", "recommender_type",
    "recommender", "tech_direction", "vendor_cert", "bio", "capable_courses",
    "teacher_cert", "note",
]

DIMENSIONS = {
    "source": "来源维度",
    "title_dim": "职称维度",
    "level": "级别维度",
    "education": "学历",
    "gender": "性别",
    "cert_category": "厂商认证",
    "availability": "可用性",
    "fee_level": "费用等级",
    "recommender_type": "推荐方类型",
    "recommender": "推荐方",
}

# 师资表中的实训室别名 → 课程体系标准名称
LAB_ALIASES: dict[str, str] = {
    "AI4S实训室": "AI for science实训室",
    "空天地一体实训室[虚仿实训室]": "空天地一体实训室",
}


def normalize_lab_name(name: str, canonical_labs: list[str] | None = None) -> str:
    """将师资表中的实训室名称统一为课程体系标准名称。"""
    name = name.strip()
    if not name:
        return name
    if name in LAB_ALIASES:
        name = LAB_ALIASES[name]
    if not canonical_labs:
        return name
    if name in canonical_labs:
        return name
    simplified = re.sub(r"\[.*?\]", "", name).strip()
    if simplified in canonical_labs:
        return simplified
    for lab in canonical_labs:
        if lab == simplified or lab.startswith(simplified) or simplified in lab:
            return lab
    return name


def _normalize_teacher_labs(teachers: list[dict], canonical_labs: list[str] | None) -> None:
    for t in teachers:
        raw_labs = _split_multi(t.get("matched_labs", ""))
        if not raw_labs and t.get("lab_list"):
            raw_labs = list(t["lab_list"])
        normalized: list[str] = []
        seen: set[str] = set()
        for lab in raw_labs:
            n = normalize_lab_name(lab, canonical_labs)
            if n and n not in seen:
                seen.add(n)
                normalized.append(n)
        t["lab_list"] = normalized
        t["matched_labs"] = "\n".join(normalized)
        t["matched_lab_count"] = len(normalized)


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _parse_priority_score(text: str) -> int | None:
    m = re.search(r"(\d+)\s*分", text)
    return int(m.group(1)) if m else None


def _parse_priority_label(text: str) -> str:
    if not text:
        return "待评估"
    if "强烈推荐" in text:
        return "强烈推荐"
    if "可考虑" in text:
        return "可考虑"
    if "待评估" in text:
        return "待评估"
    return text.split("-")[-1].strip() if "-" in text else text


def _split_multi(text: str) -> list[str]:
    if not text:
        return []
    parts = re.split(r"[、,，/;；\n]+", text)
    return [p.strip() for p in parts if p.strip()]


def _simplify_experience(text: str) -> str:
    if not text:
        return "未知"
    head = text.split("|")[0].strip()
    if "（" in head:
        head = head.split("（")[0].strip()
    return head or "未知"


def _find_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        cells = [_cell_str(c) for c in row]
        if "姓名" in cells and "来源维度" in "".join(cells):
            return i
    return HEADER_ROW


def _parse_main_sheet(ws) -> list[dict]:
    header_row = _find_header_row(ws)
    teachers: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not _cell_str(row[1] if len(row) > 1 else ""):
            continue
        record: dict[str, Any] = {"id": len(teachers) + 1}
        for idx, key in enumerate(MAIN_COLUMNS):
            val = _cell_str(row[idx]) if idx < len(row) else ""
            record[key] = val
        try:
            record["matched_lab_count"] = int(float(record["matched_lab_count"] or 0))
        except (ValueError, TypeError):
            record["matched_lab_count"] = len(_split_multi(record["matched_labs"]))
        record["priority_score"] = _parse_priority_score(record["priority"])
        record["priority_label"] = _parse_priority_label(record["priority"])
        record["experience_simple"] = _simplify_experience(record["experience"])
        record["field_tags"] = _split_multi(record["field"])
        record["lab_list"] = _split_multi(record["matched_labs"])
        teachers.append(record)
    return teachers


def _count_distribution(items: list[str]) -> list[dict]:
    counter = Counter(items)
    total = sum(counter.values()) or 1
    result = [{"label": k, "count": v, "ratio": round(v / total, 4)} for k, v in counter.most_common()]
    return result


def _build_field_stats(teachers: list[dict]) -> list[dict]:
    tags: list[str] = []
    for t in teachers:
        tags.extend(t.get("field_tags") or [])
    return _count_distribution(tags)


def _build_lab_stats(teachers: list[dict], canonical_labs: list[str] | None = None) -> list[dict]:
    counter: Counter[str] = Counter()
    for t in teachers:
        for lab in t.get("lab_list") or []:
            counter[lab] += 1
    total = len(teachers) or 1
    if canonical_labs:
        items = [(lab, counter.get(lab, 0)) for lab in canonical_labs]
        items.sort(key=lambda x: (-x[1], x[0]))
    else:
        items = counter.most_common()
    return [
        {"lab": lab, "count": cnt, "ratio": round(cnt / total, 4)}
        for lab, cnt in items
    ]


def _build_priority_stats(teachers: list[dict]) -> list[dict]:
    buckets = {"80+": 0, "60-79": 0, "40-59": 0, "20-39": 0, "0-19": 0, "未评分": 0}
    for t in teachers:
        score = t.get("priority_score")
        if score is None:
            buckets["未评分"] += 1
        elif score >= 80:
            buckets["80+"] += 1
        elif score >= 60:
            buckets["60-79"] += 1
        elif score >= 40:
            buckets["40-59"] += 1
        elif score >= 20:
            buckets["20-39"] += 1
        else:
            buckets["0-19"] += 1
    total = len(teachers) or 1
    order = ["80+", "60-79", "40-59", "20-39", "0-19", "未评分"]
    return [{"label": k, "count": buckets[k], "ratio": round(buckets[k] / total, 4)} for k in order if buckets[k]]


def _parse_table_sheet(ws, header_keywords: tuple[str, ...]) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    headers: list[str] = []
    for i, row in enumerate(rows):
        cells = [_cell_str(c) for c in row]
        if any(kw in cells for kw in header_keywords):
            header_idx = i
            headers = cells
            break
    if header_idx is None:
        return []
    result = []
    for row in rows[header_idx + 1 :]:
        cells = [_cell_str(c) for c in row]
        if not any(cells):
            continue
        if cells[0] in ("合计", "总计", ""):
            continue
        item = {headers[j]: cells[j] if j < len(cells) else "" for j in range(len(headers)) if headers[j]}
        if not item.get(headers[0]):
            continue
        result.append(item)
    return result


def _parse_top_priority(ws) -> list[dict]:
    items = []
    started = False
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_str(c) for c in row]
        if not started:
            if cells[0] == "排名" or (len(cells) > 1 and cells[1] == "姓名"):
                started = True
            continue
        if not cells[1]:
            continue
        items.append({
            "rank": int(float(cells[0])) if cells[0].replace(".", "").isdigit() else len(items) + 1,
            "name": cells[1],
            "priority": cells[2] if len(cells) > 2 else "",
            "title_dim": cells[3] if len(cells) > 3 else "",
            "source": cells[4] if len(cells) > 4 else "",
            "recommender_type": cells[5] if len(cells) > 5 else "",
            "matched_labs": cells[6] if len(cells) > 6 else "",
            "keywords": cells[7] if len(cells) > 7 else "",
            "priority_score": _parse_priority_score(cells[2] if len(cells) > 2 else ""),
        })
    if items:
        return items
    # fallback from teachers
    return []


def _parse_lab_recommender_matrix(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        cells = [_cell_str(c) for c in row]
        if cells and "实训室名称" in cells:
            header_idx = i
            break
    if header_idx is None:
        return {"headers": [], "rows": []}
    headers = [_cell_str(c) for c in rows[header_idx] if _cell_str(c)]
    data_rows = []
    for row in rows[header_idx + 1 :]:
        cells = [_cell_str(c) for c in row]
        if not cells[0] or cells[0] == "合计":
            continue
        data_rows.append(cells[: len(headers)])
    return {"headers": headers, "rows": data_rows}


def _build_dimension_stats(teachers: list[dict], canonical_labs: list[str] | None = None) -> dict[str, list[dict]]:
    stats: dict[str, list[dict]] = {}
    for key in DIMENSIONS:
        if key == "recommender":
            values = [t.get("recommender") or "未知" for t in teachers]
        else:
            values = [t.get(key) or "未知" for t in teachers]
        stats[key] = _count_distribution(values)
    stats["experience"] = _count_distribution([t.get("experience_simple") or "未知" for t in teachers])
    stats["field"] = _build_field_stats(teachers)
    stats["priority_bucket"] = _build_priority_stats(teachers)
    stats["lab_match"] = _build_lab_stats(teachers, canonical_labs)
    return stats


def _build_top_from_teachers(teachers: list[dict], limit: int = 20) -> list[dict]:
    ranked = sorted(
        teachers,
        key=lambda t: (t.get("priority_score") is not None, t.get("priority_score") or 0),
        reverse=True,
    )
    result = []
    for i, t in enumerate(ranked[:limit], start=1):
        result.append({
            "rank": i,
            "name": t["name"],
            "priority": t["priority"],
            "title_dim": t["title_dim"],
            "source": t["source"],
            "recommender_type": t["recommender_type"],
            "matched_labs": t["matched_labs"],
            "keywords": t["keywords"],
            "priority_score": t.get("priority_score"),
        })
    return result


def parse_faculty_workbook(filepath: str, canonical_labs: list[str] | None = None) -> dict:
    path = Path(filepath)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    teachers: list[dict] = []
    top_priority: list[dict] = []
    lab_recommender_matrix: dict = {"headers": [], "rows": []}
    sheet_lab_stats: list[dict] = []
    sheet_source_stats: list[dict] = []
    sheet_recommender_stats: list[dict] = []

    for name in wb.sheetnames:
        ws = wb[name]
        if "总表" in name or name == wb.sheetnames[0]:
            teachers = _parse_main_sheet(ws)
        elif "TOP" in name.upper() or "优先级" in name:
            top_priority = _parse_top_priority(ws)
        elif "交叉" in name:
            lab_recommender_matrix = _parse_lab_recommender_matrix(ws)
        elif "实训室" in name and "匹配" in name:
            for item in _parse_table_sheet(ws, ("实训室名称",)):
                lab = item.get("实训室名称", "")
                cnt = item.get("匹配师资数量", "0")
                ratio = item.get("匹配率", "0")
                if lab:
                    try:
                        count = int(float(cnt))
                    except ValueError:
                        count = 0
                    try:
                        r = float(ratio)
                    except ValueError:
                        r = 0
                    sheet_lab_stats.append({"lab": lab, "count": count, "ratio": r})
        elif "来源" in name:
            for item in _parse_table_sheet(ws, ("来源类别",)):
                label = item.get("来源类别", "")
                if label:
                    try:
                        count = int(float(item.get("人数", 0)))
                    except ValueError:
                        count = 0
                    try:
                        ratio = float(item.get("占比", 0))
                    except ValueError:
                        ratio = 0
                    sheet_source_stats.append({
                        "label": label,
                        "count": count,
                        "ratio": ratio,
                        "detail": item.get("主要来源/推荐方", ""),
                    })
        elif "推荐方" in name and "交叉" not in name:
            for item in _parse_table_sheet(ws, ("推荐方",)):
                label = item.get("推荐方", "")
                if label:
                    try:
                        count = int(float(item.get("人数", 0)))
                    except ValueError:
                        count = 0
                    try:
                        ratio = float(item.get("占比", 0))
                    except ValueError:
                        ratio = 0
                    sheet_recommender_stats.append({
                        "label": label,
                        "count": count,
                        "ratio": ratio,
                        "type": item.get("推荐方类型", ""),
                        "detail": item.get("主要师资方向", ""),
                    })

    wb.close()

    if not teachers:
        raise ValueError("未找到师资数据，请确认 Excel 包含「师资多维分析总表」")

    _normalize_teacher_labs(teachers, canonical_labs)

    if canonical_labs:
        for item in sheet_lab_stats:
            item["lab"] = normalize_lab_name(item["lab"], canonical_labs)
        if lab_recommender_matrix.get("rows"):
            for row in lab_recommender_matrix["rows"]:
                if row:
                    row[0] = normalize_lab_name(row[0], canonical_labs)

    stats = _build_dimension_stats(teachers, canonical_labs)
    if sheet_lab_stats:
        stats["lab_match_sheet"] = sheet_lab_stats
    if sheet_source_stats:
        stats["source_sheet"] = sheet_source_stats
    if sheet_recommender_stats:
        stats["recommender_sheet"] = sheet_recommender_stats

    if not top_priority:
        top_priority = _build_top_from_teachers(teachers)

    lab_set = set()
    for t in teachers:
        lab_set.update(t.get("lab_list") or [])

    lab_total = len(canonical_labs) if canonical_labs else len(lab_set)

    return {
        "meta": {
            "total_teachers": len(teachers),
            "total_labs": lab_total,
            "source_file": path.name,
            "canonical_labs": canonical_labs or sorted(lab_set),
        },
        "teachers": teachers,
        "stats": stats,
        "dimensions": DIMENSIONS,
        "top_priority": top_priority,
        "lab_recommender_matrix": lab_recommender_matrix,
    }
