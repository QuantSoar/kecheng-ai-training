# -*- coding: utf-8 -*-
"""生成课程 / 师资 Excel 上传模板（与 frontend/src/templates/spec.ts 保持同步）"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "frontend" / "public" / "templates"

COURSE_TYPES = ["通识课程", "专业基础课", "专业核心课", "专业实践课"]

COURSE_DETAIL_HEADERS = [
    "序号", "实训室名称", "课程类型", "课程名称", "建议学时",
    "推荐教材", "实训平台", "课程来源", "备注",
]

VENDOR_HEADERS = [
    "序号", "归属实训室（标准）", "原始实训室名称", "课程名称", "课程描述",
    "课程类型（标准）", "原始课程类型", "是否绑定硬件", "配套硬件",
    "参考课时（总）", "参考课时（理论）", "参考课时（实训）",
    "所属厂商", "培养能力", "岗位辐射",
]

OVERVIEW_HEADERS = [
    "序号", "实训室名称", "通识课程", "专业基础课", "专业核心课", "专业实践课", "课程总数", "备注",
]

LAB_NAME_HEADERS = ["序号", "实训室名称", "原Excel中名称"]

STANDARD_LABS = [
    "人工智能应用开发实训室", "机器视觉和VLM实训室", "智能边缘计算应用开发实训室",
    "视频语音智能体实训室", "鸿蒙智联实训室", "低空智能实训室", "具身智能实训室",
    "电子数据调查分析实训室", "AI+网络空间安全实训室", "互联网舆情监测与管控实验室",
    "智能体实训室", "高性能智算实训室", "AIGC视觉传达实训室", "AI for science实训室",
    "空天地一体实训室", "洁净实验室", "集成电路与电子信息实训室",
    "TPU异构算力与软硬协同实训室", "卫星互联网实训室", "AI高端装备制造实训室",
    "智能交通行业应用开发实训室", "智能制造行业开发实训室",
    "电子数据取证司法鉴定实训室", "数据标注实训室",
]

FACULTY_HEADERS = [
    "序号", "姓名", "性别", "学历", "毕业院校", "专业", "职级",
    "来源维度", "职称维度", "级别维度", "经验维度", "厂商认证", "可用性",
    "核心能力", "专业领域", "关键词", "匹配实训室", "匹配实训室数",
    "教学管理经验", "合作模式", "费用等级", "推荐优先级",
    "推荐方类型", "推荐方", "技术方向", "厂商认证证书", "讲师简介", "可授课程",
    "教师资格证", "备注",
]

TITLE_FILL = PatternFill("solid", fgColor="FFF3E8")
HEADER_FILL = PatternFill("solid", fgColor="FFE4CC")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, row: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _autowidth(ws, max_col: int, min_width: int = 10, max_width: int = 36) -> None:
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = min(max_width, max(min_width, 12))


def _write_guide(ws, lines: list[str]) -> None:
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = WRAP
        if i == 1:
            cell.font = Font(bold=True, size=14)


def generate_course_template() -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("填写说明", 0)
    _write_guide(guide, [
        "AI 课程图谱 — 课程 Excel 上传模板",
        "",
        "【必填】全部课程明细 — 驱动总览、实训室、课程检索、共享网络、大屏等全部课程相关页面",
        "【建议填】各个厂商课程体系 — 驱动厂商课程页的长描述、硬件、课时等详情（与明细按「实训室+课程名」关联）",
        "【可选】课程体系总览 / 实训室名称 / 统计汇总 — 不填也可，系统会从「全部课程明细」自动汇总",
        "",
        "填写要点：",
        "1. 课程类型只能填：通识课程 / 专业基础课 / 专业核心课 / 专业实践课",
        "2. 实训室名称请使用「实训室名称」Sheet 中的标准名称，保证与师资表一致",
        "3. 厂商课在「课程来源」中标注，如：厂商课程/华为、厂商课程/华清智言（「华清」会自动归一为华清智言）",
        "4. 每个实训室可在明细中插入分组行，如：「1  人工智能应用开发实训室」（仅 A 列有文字，其余留空）",
        "5. 填好后保存为 .xlsx，在系统左侧点击「上传课程 Excel」",
    ])

    ws = wb.create_sheet("全部课程明细")
    ws.merge_cells("A1:I1")
    ws["A1"] = "24个实训室课程体系建设 — 全部课程明细"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:I2")
    ws["A2"] = "课程类型：通识课程 / 专业基础课 / 专业核心课 / 专业实践课  |  厂商课请在「课程来源」标注"
    for i, h in enumerate(COURSE_DETAIL_HEADERS, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(COURSE_DETAIL_HEADERS))
    ws.cell(row=5, column=1, value="1  人工智能应用开发实训室")
    examples = [
        (1, "人工智能应用开发实训室", "专业基础课", "操作系统原理", 48,
         "《计算机操作系统》(第4版) 汤小丹", "Linux虚拟机实验环境", "高校核心课", ""),
        (2, "人工智能应用开发实训室", "专业实践课", "昇腾AI行业应用开发综合实训", 64,
         "《昇腾AI应用开发实战》", "华为Atlas 500 Pro", "厂商课程/华为", "厂商课示例"),
    ]
    for r, row in enumerate(examples, start=6):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _autowidth(ws, len(COURSE_DETAIL_HEADERS))

    ws_v = wb.create_sheet("各个厂商课程体系")
    ws_v.merge_cells("A1:O1")
    ws_v["A1"] = "厂商课程详情（与全部课程明细按 实训室+课程名 关联）"
    ws_v["A1"].font = Font(bold=True)
    for i, h in enumerate(VENDOR_HEADERS, start=1):
        ws_v.cell(row=3, column=i, value=h)
    _style_header_row(ws_v, 3, len(VENDOR_HEADERS))
    vendor_example = [
        1, "人工智能应用开发实训室", "", "昇腾AI行业应用开发综合实训",
        "面向昇腾平台的行业 AI 应用开发综合实训，含模型训练与部署全流程。",
        "专业实践课", "", "是", "华为Atlas 500 Pro",
        64, 16, 48, "华为", "模型训练与部署", "AI应用开发工程师",
    ]
    for c, val in enumerate(vendor_example, start=1):
        ws_v.cell(row=4, column=c, value=val)
    _autowidth(ws_v, len(VENDOR_HEADERS))

    ws_o = wb.create_sheet("课程体系总览")
    ws_o.merge_cells("A1:H1")
    ws_o["A1"] = "课程体系总览（可选，系统也会从明细自动统计）"
    for i, h in enumerate(OVERVIEW_HEADERS, start=1):
        ws_o.cell(row=4, column=i, value=h)
    _style_header_row(ws_o, 4, len(OVERVIEW_HEADERS))
    ws_o.cell(row=5, column=1, value="（示例，请按实际填写或留空）")
    _autowidth(ws_o, len(OVERVIEW_HEADERS))

    ws_l = wb.create_sheet("实训室名称")
    ws_l.merge_cells("A1:C1")
    ws_l["A1"] = "标准实训室名称对照表（建议与明细、师资表保持一致）"
    for i, h in enumerate(LAB_NAME_HEADERS, start=1):
        ws_l.cell(row=3, column=i, value=h)
    _style_header_row(ws_l, 3, len(LAB_NAME_HEADERS))
    for i, lab in enumerate(STANDARD_LABS, start=1):
        ws_l.cell(row=i + 3, column=1, value=i)
        ws_l.cell(row=i + 3, column=2, value=lab)
        ws_l.cell(row=i + 3, column=3, value="")
    _autowidth(ws_l, 3)

    ws_s = wb.create_sheet("统计汇总")
    ws_s["A1"] = "统计汇总（可选，系统会从明细重新计算，此处可留空）"
    ws_s["A3"] = "课程类型分布"
    ws_s["A4"] = "类型"
    ws_s["B4"] = "数量"
    ws_s["C4"] = "占比"
    _autowidth(ws_s, 3)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "课程图谱上传模板.xlsx"
    wb.save(path)
    return path


def generate_faculty_template() -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("填写说明", 0)
    _write_guide(guide, [
        "AI 课程图谱 — 师资 Excel 上传模板",
        "",
        "【必填】师资多维分析总表 — 驱动师资分析、综合匹配、大屏师资章节",
        "【可选】其余 Sheet — 用于补充 TOP 列表与统计展示，不填则系统从总表自动计算",
        "",
        "填写要点：",
        "1. 表头必须在第 4 行，且包含「姓名」「来源维度」等列（本模板已预设，请勿删改列顺序）",
        "2. 「匹配实训室」多个时用顿号、逗号或换行分隔，名称请与课程 Excel 中实训室名称一致",
        "3. 「推荐优先级」建议格式：「85分-强烈推荐」，系统会自动提取分数与标签",
        "4. 「级别维度」填：高 / 中 / 低",
        "5. 先上传课程 Excel 再上传师资 Excel，可自动对齐 24 个实训室标准名称",
        "6. 填好后保存为 .xlsx，在系统左侧点击「上传师资 Excel」",
    ])

    ws = wb.create_sheet("师资多维分析总表")
    ws.merge_cells("A1:AD1")
    ws["A1"] = "师资多维分析总表"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:AD2")
    ws["A2"] = "第 4 行为表头，第 5 行起填数据；「匹配实训室」须与课程体系标准名称一致"
    for i, h in enumerate(FACULTY_HEADERS, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(FACULTY_HEADERS))
    example = [
        1, "张三", "男", "博士", "某某大学", "计算机科学", "副教授",
        "高校", "副教授", "高", "10年+高校教学 | 企业兼职",
        "华为认证", "可排课", "深度学习、模型部署", "人工智能、计算机视觉",
        "PyTorch,昇腾,模型部署", "人工智能应用开发实训室、具身智能实训室", 2,
        "有", "兼职/项目", "中",
        "85分-强烈推荐", "高校", "某某大学", "昇腾AI", "HCIA-AI",
        "长期从事 AI 教学与项目指导……", "操作系统原理、AI模型部署", "有", "",
    ]
    for c, val in enumerate(example, start=1):
        ws.cell(row=5, column=c, value=val)
    _autowidth(ws, len(FACULTY_HEADERS), min_width=8, max_width=28)

    ws_top = wb.create_sheet("TOP20推荐优先级")
    ws_top["A1"] = "TOP20 推荐优先级（可选，不填则从总表按分数自动生成）"
    top_headers = ["排名", "姓名", "推荐优先级", "职称维度", "来源维度", "推荐方类型", "匹配实训室", "关键词"]
    for i, h in enumerate(top_headers, start=1):
        ws_top.cell(row=3, column=i, value=h)
    _style_header_row(ws_top, 3, len(top_headers))
    _autowidth(ws_top, len(top_headers))

    ws_lab = wb.create_sheet("实训室师资匹配")
    ws_lab["A1"] = "实训室师资匹配统计（可选）"
    for i, h in enumerate(["实训室名称", "匹配师资数量", "匹配率"], start=1):
        ws_lab.cell(row=3, column=i, value=h)
    _style_header_row(ws_lab, 3, 3)
    _autowidth(ws_lab, 3)

    ws_src = wb.create_sheet("来源维度统计")
    ws_src["A1"] = "来源维度统计（可选）"
    for i, h in enumerate(["来源类别", "人数", "占比", "主要来源/推荐方"], start=1):
        ws_src.cell(row=3, column=i, value=h)
    _style_header_row(ws_src, 3, 4)
    _autowidth(ws_src, 4)

    ws_rec = wb.create_sheet("推荐方统计")
    ws_rec["A1"] = "推荐方统计（可选）"
    for i, h in enumerate(["推荐方", "人数", "占比", "推荐方类型", "主要师资方向"], start=1):
        ws_rec.cell(row=3, column=i, value=h)
    _style_header_row(ws_rec, 3, 5)
    _autowidth(ws_rec, 5)

    ws_mx = wb.create_sheet("实训室推荐方交叉")
    ws_mx["A1"] = "实训室 × 推荐方交叉矩阵（可选）"
    ws_mx.cell(row=3, column=1, value="实训室名称")
    _autowidth(ws_mx, 3)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "师资多维分析上传模板.xlsx"
    wb.save(path)
    return path


    print(f"已生成: {f}")


def generate_job_template() -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("填写说明", 0)
    _write_guide(guide, [
        "AI 课程图谱 — 岗位映射 Excel 上传模板",
        "",
        "【必填】岗位-实训室映射矩阵 — 驱动岗位列表与实训室关联",
        "【必填】岗位能力分层拆解 — 驱动岗位技能图谱展示",
        "【建议填】实训室-岗位-课程三对照 — 驱动课程/师资联动区块",
        "【可选】信息缺口与补充建议",
        "",
        "填写要点：",
        "1. 映射矩阵实训室列填：● 核心 / ◐ 重要 / ○ 基础，留空表示无关联",
        "2. 「核心课程」「推荐课程」多个时用顿号、逗号分隔，须与课程 Excel 名称尽量一致",
        "3. 填好后保存为 .xlsx，在系统「数据管理」页上传",
    ])

    matrix_headers = [
        "序号", "岗位大类", "岗位名称", "核心关键词", "市场热度",
        "人工智能应用开发", "高性能智算", "智能体", "数据标注",
    ]
    ws_m = wb.create_sheet("岗位-实训室映射矩阵")
    ws_m.merge_cells("A1:I1")
    ws_m["A1"] = "岗位-实训室映射矩阵"
    ws_m["A1"].font = Font(bold=True, size=12)
    ws_m.merge_cells("A2:I2")
    ws_m["A2"] = "实训室列使用简称；关联符号：● 核心 / ◐ 重要 / ○ 基础"
    for i, h in enumerate(matrix_headers, start=1):
        ws_m.cell(row=4, column=i, value=h)
    _style_header_row(ws_m, 4, len(matrix_headers))
    examples = [
        [1, "基础设施类", "智算中心运维工程师", "Linux,GPU,集群", "高", "○", "●", "◐", ""],
        [2, "核心技术类", "AI平台工程师", "K8s,容器,调度", "高", "●", "◐", "●", "○"],
    ]
    for r, row in enumerate(examples, start=5):
        for c, val in enumerate(row, start=1):
            ws_m.cell(row=r, column=c, value=val)
    _autowidth(ws_m, len(matrix_headers))

    skill_headers = [
        "序号", "岗位名称", "岗位大类", "能力域", "初级技能点", "中级技能点",
        "高级技能点", "关联实训室", "推荐课程",
    ]
    ws_s = wb.create_sheet("岗位能力分层拆解")
    ws_s.merge_cells("A1:I1")
    ws_s["A1"] = "岗位能力分层拆解"
    for i, h in enumerate(skill_headers, start=1):
        ws_s.cell(row=4, column=i, value=h)
    _style_header_row(ws_s, 4, len(skill_headers))
    skill_examples = [
        [1, "智算中心运维工程师", "基础设施类", "系统运维能力",
         "Linux基础命令与系统管理", "集群监控与容量规划", "架构设计与自动化运维",
         "高性能智算实训室", "Linux系统管理、云计算基础"],
        [2, "智算中心运维工程师", "基础设施类", "故障排查与优化",
         "硬件故障定位", "性能调优与日志分析", "智能运维体系建设",
         "高性能智算实训室", "AI超算集群搭建与运维实训"],
    ]
    for r, row in enumerate(skill_examples, start=5):
        for c, val in enumerate(row, start=1):
            ws_s.cell(row=r, column=c, value=val)
    _autowidth(ws_s, len(skill_headers), min_width=8, max_width=28)

    lc_headers = [
        "序号", "实训室名称", "能力域分组", "辐射岗位", "核心课程",
        "课程类型", "关联等级", "培养能力", "岗位大类",
    ]
    ws_lc = wb.create_sheet("实训室-岗位-课程三对照")
    ws_lc.merge_cells("A1:I1")
    ws_lc["A1"] = "实训室-岗位-课程三对照"
    for i, h in enumerate(lc_headers, start=1):
        ws_lc.cell(row=4, column=i, value=h)
    _style_header_row(ws_lc, 4, len(lc_headers))
    lc_example = [
        1, "高性能智算实训室", "系统运维", "智算中心运维工程师",
        "AI超算集群搭建与运维实训、智算中心资源调度与优化实训",
        "专业实践课", "●", "Linux运维/GPU集群/分布式存储", "基础设施类",
    ]
    for c, val in enumerate(lc_example, start=1):
        ws_lc.cell(row=5, column=c, value=val)
    _autowidth(ws_lc, len(lc_headers), min_width=8, max_width=28)

    gap_headers = ["序号", "缺口类别", "缺口描述", "影响范围", "严重程度", "补充建议"]
    ws_g = wb.create_sheet("信息缺口与补充建议")
    ws_g.merge_cells("A1:F1")
    ws_g["A1"] = "信息缺口与补充建议（可选）"
    for i, h in enumerate(gap_headers, start=1):
        ws_g.cell(row=4, column=i, value=h)
    _style_header_row(ws_g, 4, len(gap_headers))
    ws_g.cell(row=5, column=1, value=1)
    ws_g.cell(row=5, column=2, value="能力域覆盖")
    ws_g.cell(row=5, column=3, value="部分岗位能力域行数较少")
    ws_g.cell(row=5, column=4, value="岗位图谱展示")
    ws_g.cell(row=5, column=5, value="中")
    ws_g.cell(row=5, column=6, value="在「岗位能力分层拆解」中为岗位增加多行能力域")
    _autowidth(ws_g, len(gap_headers), min_width=8, max_width=32)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "岗位映射上传模板.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    c = generate_course_template()
    f = generate_faculty_template()
    j = generate_job_template()
    print(f"已生成: {c}")
    print(f"已生成: {f}")
    print(f"已生成: {j}")
