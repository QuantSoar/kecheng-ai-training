# -*- coding: utf-8 -*-
"""解析岗位映射 Excel（与 frontend/src/parsers/jobMapParser.ts 逻辑一致）"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

STANDARD_LABS = [
    "人工智能应用开发实训室", "机器视觉和VLM实训室", "智能边缘计算应用开发实训室",
    "视频语音智能体实训室", "鸿蒙智联实训室", "低空智能实训室", "具身智能实训室",
    "电子数据调查分析实训室", "AI+网络空间安全实训室", "互联网舆情监测与管控实验室",
    "智能体实训室", "高性能智算实训室", "AIGC视觉传达实训室", "AI for science实训室",
    "空天地一体实训室", "洁净实验室", "集成电路与电子信息实训室",
    "TPU异构算力与软硬协同实训室", "卫星互联网实训室", "AI高端装备制造实训室",
    "智能交通行业应用开发实训室", "智能制造行业开发实训室",
    "电子数据取证司法鉴定实训室", "数据标注实训室",
]

LINK_MAP = {"●": "core", "◐": "important", "○": "basic"}


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _split_list(text: str) -> list[str]:
    if not text:
        return []
    return [p.strip() for p in re.split(r"[,，、/;；\n]+", text) if p.strip()]


def _parse_link(val: Any) -> str:
    s = _cell_str(val)
    return LINK_MAP.get(s, "none")


def normalize_matrix_lab_name(short: str) -> str:
    s = short.strip()
    if not s:
        return s
    for lab in STANDARD_LABS:
        if lab == s or lab.startswith(s):
            return lab
        base = lab.replace("实训室", "").replace("实验室", "")
        if s == base or s in lab:
            return lab
    if not s.endswith("实训室") and not s.endswith("实验室"):
        return f"{s}实训室"
    return s


def _sheet_rows(wb: openpyxl.Workbook, name: str) -> list[list[Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row) if row else [])
    return rows


def _find_sheet(wb: openpyxl.Workbook, test: str) -> str | None:
    for n in wb.sheetnames:
        if test in n:
            return n
    return None


def _find_header_row(rows: list[list[Any]], markers: list[str]) -> int:
    for i, row in enumerate(rows):
        text = " ".join(_cell_str(c) for c in row)
        if all(m in text for m in markers):
            return i
    return 0


def _parse_matrix_sheet(wb: openpyxl.Workbook) -> tuple[dict[str, dict], list[str]]:
    sheet = _find_sheet(wb, "映射矩阵")
    if not sheet:
        return {}, []
    rows = _sheet_rows(wb, sheet)
    header_idx = _find_header_row(rows, ["岗位名称", "岗位大类"])
    headers = [_cell_str(c) for c in rows[header_idx]]
    lab_start = next(
        (i for i, h in enumerate(headers) if "应用开发" in h or "智算" in h or h == "人工智能应用开发"),
        -1,
    )
    start_col = lab_start if lab_start >= 5 else 5
    lab_columns = [h for h in headers[start_col:] if h]
    jobs: dict[str, dict] = {}

    for row in rows[header_idx + 1 :]:
        name = _cell_str(row[2] if len(row) > 2 else "")
        if not name:
            continue
        lab_links = []
        for i, lab in enumerate(lab_columns):
            col = start_col + i
            level = _parse_link(row[col] if col < len(row) else "")
            if level != "none":
                lab_links.append({"lab": normalize_matrix_lab_name(lab), "level": level})
        jobs[name] = {
            "id": int(_cell_str(row[0])) if len(row) > 0 and _cell_str(row[0]).isdigit() else len(jobs) + 1,
            "name": name,
            "category": _cell_str(row[1] if len(row) > 1 else ""),
            "keywords": _cell_str(row[3] if len(row) > 3 else ""),
            "heat": _cell_str(row[4] if len(row) > 4 else ""),
            "skills": [],
            "labLinks": lab_links,
        }
    return jobs, lab_columns


def _parse_skills_sheet(wb: openpyxl.Workbook, jobs: dict[str, dict]) -> None:
    sheet = _find_sheet(wb, "能力分层")
    if not sheet:
        return
    rows = _sheet_rows(wb, sheet)
    header_idx = _find_header_row(rows, ["岗位名称", "能力域"])
    for row in rows[header_idx + 1 :]:
        name = _cell_str(row[1] if len(row) > 1 else "")
        if not name:
            continue
        skill = {
            "domain": _cell_str(row[3] if len(row) > 3 else ""),
            "junior": _cell_str(row[4] if len(row) > 4 else ""),
            "mid": _cell_str(row[5] if len(row) > 5 else ""),
            "senior": _cell_str(row[6] if len(row) > 6 else ""),
            "labs": _split_list(_cell_str(row[7] if len(row) > 7 else "")),
            "courses": _split_list(_cell_str(row[8] if len(row) > 8 else "")),
        }
        if not skill["domain"]:
            continue
        if name in jobs:
            jobs[name]["skills"].append(skill)
            if not jobs[name]["category"]:
                jobs[name]["category"] = _cell_str(row[2] if len(row) > 2 else "")
        else:
            jobs[name] = {
                "id": int(_cell_str(row[0])) if _cell_str(row[0]).isdigit() else len(jobs) + 1,
                "name": name,
                "category": _cell_str(row[2] if len(row) > 2 else ""),
                "keywords": "",
                "heat": "",
                "skills": [skill],
                "labLinks": [],
            }


def _parse_lab_course_sheet(wb: openpyxl.Workbook) -> list[dict]:
    sheet = _find_sheet(wb, "课程对照") or _find_sheet(wb, "三对照")
    if not sheet:
        return []
    rows = _sheet_rows(wb, sheet)
    header_idx = _find_header_row(rows, ["实训室名称", "辐射岗位"])
    result: list[dict] = []
    for row in rows[header_idx + 1 :]:
        lab = _cell_str(row[1] if len(row) > 1 else "")
        job = _cell_str(row[3] if len(row) > 3 else "")
        if not lab or not job:
            continue
        result.append({
            "lab": lab,
            "domainGroup": _cell_str(row[2] if len(row) > 2 else ""),
            "job": job,
            "courses": _split_list(_cell_str(row[4] if len(row) > 4 else "")),
            "courseType": _cell_str(row[5] if len(row) > 5 else ""),
            "linkLevel": _parse_link(row[6] if len(row) > 6 else ""),
            "capability": _cell_str(row[7] if len(row) > 7 else ""),
            "jobCategory": _cell_str(row[8] if len(row) > 8 else ""),
        })
    return result


def _parse_gaps_sheet(wb: openpyxl.Workbook) -> list[dict]:
    sheet = _find_sheet(wb, "信息缺口") or _find_sheet(wb, "补充建议")
    if not sheet:
        return []
    rows = _sheet_rows(wb, sheet)
    header_idx = _find_header_row(rows, ["缺口类别", "缺口描述"])
    result: list[dict] = []
    for row in rows[header_idx + 1 :]:
        category = _cell_str(row[1] if len(row) > 1 else "")
        description = _cell_str(row[2] if len(row) > 2 else "")
        if not category or not description:
            continue
        result.append({
            "category": category,
            "description": description,
            "scope": _cell_str(row[3] if len(row) > 3 else ""),
            "severity": _cell_str(row[4] if len(row) > 4 else ""),
            "suggestion": _cell_str(row[5] if len(row) > 5 else ""),
        })
    return result


def parse_job_map_workbook(filepath: str | Path) -> dict:
    path = Path(filepath)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        jobs_map, lab_columns = _parse_matrix_sheet(wb)
        _parse_skills_sheet(wb, jobs_map)
        job_list = sorted(jobs_map.values(), key=lambda j: j["id"])
        return {
            "meta": {
                "total_jobs": len(job_list),
                "total_skills": sum(len(j["skills"]) for j in job_list),
                "total_lab_links": sum(len(j["labLinks"]) for j in job_list),
                "source_file": path.name,
            },
            "jobs": job_list,
            "labCourseRows": _parse_lab_course_sheet(wb),
            "gaps": _parse_gaps_sheet(wb),
            "labColumns": lab_columns,
        }
    finally:
        wb.close()
